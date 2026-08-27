"""tests/e2e/e2e_loader.py — testcases.xlsx 加载器（收集期）。

按表头名读 testcases.xlsx（容错列重排），逐行解析为 Case：
  - dcat 命令：动词锚定正则 dcat (inject|clean|query|list|serve) 提取，
    去 CJK 后 shlex → argv（argv[0]=='dcat'，运行时由 fixture 换成绝对二进制路径）。
  - setup：前置条件含「已注入」且步骤无 inject → 解析 module+--args 构造 setup 注入；
    解析不出 → skip_reason。
  - skip_reason：无 dcat 命令 / setup 不可解析 → 预置 skip。
parametrized_cases() 产出 pytest.param（id=TC-xxx_模块 + markers）。
"""
import re
import shlex
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path

import openpyxl
import pytest

HERE = Path(__file__).parent
XLSX = HERE / "testcases.xlsx"

# 列表头（中文；.py 为 UTF-8，字面量可用）
H_ID = "用例编号"
H_TITLE = "用例标题"
H_PRE = "前置条件"
H_STEPS = "测试步骤"
H_DATA = "测试数据"
H_EXPECT = "预期结果"
H_PRIO = "优先级"
H_REQ = "关联需求"
H_RISK = "补充标识"
H_VCMD = "验证观测命令"
H_VASSERT = "验证断言"

# 动词锚定：只匹配 dcat <已知动词>，且命令为纯 ASCII——遇 CJK/全角即停止
# （避免「（无 --cores）」等描述性括号里的 flag 被误并入命令）
_DCMD_RE = re.compile(r'dcat\s+(?:inject|clean|query|list|serve)\b[\x20-\x7e]+')
# dcat 动词（用于判断步骤是否含 inject）
_INJECT_RE = re.compile(r'dcat\s+inject\b')
_SETUP_KEYWORDS = ("已注入", "成功注入", "已执行")
_MODULE_RE = re.compile(r'r\w+')
_FLAG_RE = re.compile(r'--[\w=,./\-]+')


# 精选轻量冒烟 P1（核心功能/状态不变量；非 NPU、非 root，可在 GHA ubuntu 跑）。
# rNET 注入硬编码 eth0 会打断 GHA runner → 暂留全量，不进轻量。
# 原选 22 条中 TC-530/569/590/591 因命令操作 rnet 被标 root，剔除 → 18 条。
_SMOKE_USERSPACE = frozenset({
    "TC-001", "TC-010", "TC-044", "TC-224", "TC-228", "TC-233",
    "TC-240", "TC-242", "TC-250", "TC-258", "TC-533", "TC-534",
    "TC-552", "TC-554", "TC-555", "TC-575", "TC-580", "TC-585",
})
SMOKE_IDS = _SMOKE_USERSPACE


@dataclass
class Case:
    id: str
    title: str
    module: str            # 原始模块（标题首段）
    module_slug: str        # ASCII slug（marker 用）
    priority: str
    req: str
    risktag: str
    precondition: str
    steps: str
    expected: str
    test_data: str
    vcmd: str
    vassert: str
    cmds: list = field(default_factory=list)   # list[list[str]] argv, argv[0]=='dcat'
    setup_argv: list = field(default_factory=None)  # list[str] argv 或 None
    skip_reason: str = ""


def _slug(name, sep="_"):
    s = re.sub(r'[^a-z0-9]+', sep, (name or "").lower()).strip(sep)
    return s or "misc"


def _req_slug(req):
    return re.sub(r'[^a-z0-9]+', '', (req or "").lower()) or "noreq"


def _ascii_only(s):
    return "".join(ch for ch in s if ord(ch) < 128)


def _parse_cmds(steps):
    """从测试步骤提取 dcat 命令 argv 列表。argv[0]=='dcat'。"""
    cmds = []
    for m in _DCMD_RE.finditer(steps or ""):
        cleaned = m.group(0).strip()
        if not cleaned:
            continue
        try:
            argv = shlex.split(cleaned)
        except ValueError:
            continue
        if argv and argv[0] == "dcat":
            cmds.append(argv)
    return cmds


def _indicates_prior_inject(pre):
    """前置条件是否指示需前序注入 setup（已注入/成功注入/已执行...注入）。"""
    if not pre or "注入" not in pre:
        return False
    return any(kw in pre for kw in _SETUP_KEYWORDS)


def _parse_setup(precondition, cmds=None, fallback_uid=None):
    """前置条件指示需前序注入 → 解析 setup inject argv。无/解析不出 → None。

    uid 解析优先级（TC-157 前置 'eth0 已注入 rNET_loss' 必须先注入 rNET_loss 占住
    root qdisc，步骤才是注入 rNET_jitter——若误用 steps 命令的 uid(jitter) 会把
    "已有 qdisc 注入失败" 用例变成 "同资源重注入 exit5"）：
      1. precondition 里的显式真实 uid（r(CPU|MEM|... )_X 形态）
      2. steps 首条 dcat inject 命令的 uid（precondition 仅描述性文本如
         '已注入 --cores' 无 uid 时 — 例 TC-597）
      3. 标题 module 兜底
    参数：优先带同 uid 的 steps inject 命令参数；否则 pre 的 --flags。"""
    if not _indicates_prior_inject(precondition):
        return None
    pre_flags = _FLAG_RE.findall(precondition)
    # 1) precondition 显式 uid（不匹配描述性中文，精准抓真实 r<模块>_<名>）
    m = re.search(r'\br(?:CPU|MEM|NET|PROC|DISK|FS|SYS|DOCKER|NPU)_[A-Za-z0-9_]+', precondition)
    if m:
        uid = m.group(0)
        # 同 uid 的 steps inject 命令参数更完整（含 --iface 等 pre 文本没有的）
        for argv in cmds or []:
            if len(argv) >= 3 and argv[0] == "dcat" and argv[1] == "inject" and argv[2] == uid:
                return ["dcat", "inject", uid] + list(argv[3:])
        return ["dcat", "inject", uid, *pre_flags]
    # 2) pre 无显式 uid：取 steps 首条 inject（例 TC-597 无 uid 但有 clean<uid>；标题兜底）
    for argv in cmds or []:
        if len(argv) >= 3 and argv[0] == "dcat" and argv[1] == "inject":
            return ["dcat", "inject", argv[2]] + list(argv[3:])
    # 3) 标题 module（ID 用）
    if fallback_uid:
        return ["dcat", "inject", fallback_uid, *pre_flags]
    return None


def _load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.worksheets[0]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers) if h}

    def g(row, name, default=""):
        i = idx.get(name)
        if i is None:
            return default
        v = ws.cell(row=row, column=i + 1).value
        return v if v is not None else default

    rows = []
    for r in range(2, ws.max_row + 1):
        if not g(r, H_ID):
            continue
        rows.append({h: g(r, h) for h in idx})
    return rows


@lru_cache(maxsize=1)
def load_cases():
    cases = []
    for row in _load_rows():
        steps = row.get(H_STEPS, "")
        cmds = _parse_cmds(steps)
        pre = row.get(H_PRE, "")
        title = row.get(H_TITLE, "")
        module_raw = title.split("-")[0] if title else ""
        setup_argv = _parse_setup(pre, cmds, fallback_uid=module_raw)
        needs_setup = _indicates_prior_inject(pre)

        skip_reason = ""
        if not cmds:
            skip_reason = "no dcat command in steps"
        elif needs_setup and not setup_argv:
            skip_reason = f"setup unparsable from precondition: {pre!r}"

        cases.append(Case(
            id=str(row.get(H_ID, "")).strip(),
            title=title,
            module=module_raw,
            module_slug=_slug(module_raw),
            priority=str(row.get(H_PRIO, "")).strip(),
            req=str(row.get(H_REQ, "")).strip(),
            risktag=str(row.get(H_RISK, "") or ""),
            precondition=pre,
            steps=steps,
            expected=row.get(H_EXPECT, ""),
            test_data=row.get(H_DATA, ""),
            vcmd=row.get(H_VCMD, ""),
            vassert=row.get(H_VASSERT, ""),
            cmds=cmds,
            setup_argv=setup_argv if (needs_setup and setup_argv) else None,
            skip_reason=skip_reason,
        ))
    return cases


def _marks_for(case):
    marks = []
    if case.priority in ("P0", "P1", "P2"):
        marks.append(getattr(pytest.mark, case.priority))
    mod = case.module.lower()
    setup_uid = ""
    if case.setup_argv and len(case.setup_argv) > 2 and case.setup_argv[1] == "inject":
        setup_uid = case.setup_argv[2].lower()
    uids = {a[2].lower() for a in case.cmds
            if len(a) > 2 and a[0] == "dcat" and a[1] in ("inject", "clean", "query")}
    uids.add(setup_uid)
    if mod.startswith("rnpu") or any(u.startswith("rnpu") for u in uids):
        marks.append(pytest.mark.hardware)
    if mod.startswith("rnet") or any(u.startswith("rnet") for u in uids):
        marks.append(pytest.mark.net)
    if (mod.startswith("rnet") or mod == "rcpu_core_offline" or mod == "rcpu_overload"
            or mod.startswith("rmem") or mod.startswith("rsys") or mod.startswith("rfs")
            or mod.startswith("rdis")
            or any(u.startswith("rnet") for u in uids)
            or any(u == "rcpu_core_offline" for u in uids)
            or any(u == "rcpu_overload" for u in uids)):
        marks.append(pytest.mark.root)
    if case.id in SMOKE_IDS:
        marks.append(pytest.mark.smoke)
    marks.append(getattr(pytest.mark, case.module_slug))
    marks.append(getattr(pytest.mark, _req_slug(case.req)))
    # xdist_group: 同故障模块的测试在同一 worker 串行, 避免网络/进程状态冲突
    mod_lower = case.module.lower()
    if mod_lower.startswith("rnet"):
        # tc qdisc 操作物理网卡的模块必须同组串行 (共用 eth0 qdisc)
        _qdisc_mods = {"rnet_bw_limit", "rnet_degrade", "rnet_jitter",
                       "rnet_reorder", "rnet_delay", "rnet_loss"}
        if mod_lower in _qdisc_mods:
            grp = "rnet_qdisc"
        elif mod_lower in ("rnet_down", "rnet_link_flap"):
            grp = "rnet_link"
        elif mod_lower == "rnet_port_occupy":
            grp = "rnet_proc"
        else:
            grp = "rnet_other"
    elif mod_lower.startswith("rproc"):
        grp = "rproc"
    elif mod_lower.startswith("rcpu"):
        grp = "rcpu"
    elif mod_lower.startswith("rmem"):
        grp = "rmem"
    elif mod_lower.startswith("rsys"):
        grp = "rsys"
    elif mod_lower.startswith("rfs"):
        grp = "rfs"
    elif mod_lower.startswith("rnpu"):
        grp = "rnpu"
    else:
        grp = "misc"
    marks.append(pytest.mark.xdist_group(grp))
    return marks


def parametrized_cases():
    return [
        pytest.param(c, id=f"{c.id}_{c.module_slug}", marks=_marks_for(c))
        for c in load_cases()
    ]
